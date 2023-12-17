from openpyxl import load_workbook
import csv
from sklearn.preprocessing import StandardScaler

map_favour_device = {
    "Mobile Phone": 0,
    "Pad": 1,
    "Phone": 2
}

map_status = {
    "Single": 0,
    "Married": 1,
    "Divorced": 2
}

map_gender = {
    "Female": 0,
    "Male": 1
}

map_favour_cat = {
    "Fashion": 0,
    "Grocery": 1,
    "Household": 2,
    "Laptop & Accessory": 3,
    "Mobile Phone": 4,
    "Others": 5
}

# 打开 Excel 文件
workbook = load_workbook('Project Dataset.xlsx')

# 获取工作表名称
sheet_names = workbook.sheetnames
print("工作表名称:", sheet_names)

# 选择要操作的工作表
sheet = workbook['E Comm']

# 读取单元格数据示例
max_row = sheet.max_row
max_column = sheet.max_column
data = []
Y = []
for row_num in range(2, max_row + 1):
    row_data = []  # 创建一个空列表来存储每一行的数据
    for col_num in range(1, max_column + 1):
        cell_value = sheet.cell(row=row_num, column=col_num).value
        if col_num == 4:
            cell_value = map_favour_device[cell_value]
        elif col_num == 7:
            cell_value = map_status[cell_value]
        elif col_num == 9:
            cell_value = map_gender[cell_value]
        elif col_num == 14:
            cell_value = map_favour_cat[cell_value]
        if cell_value is None:
            cell_value = -1
        row_data.append(cell_value)
    data.append(row_data)
    Y.append(row_data[13])
    # print(f"第{row_num}行数据:", row_data)
# print(data)
scaler = StandardScaler()
scaled_data = scaler.fit_transform(data)

A_filename = 'A.csv'
B_filename = 'B.csv'

A = open(A_filename, 'w', newline='', encoding='utf-8')
B = open(B_filename, 'w', newline='', encoding='utf-8')
A_writer = csv.writer(A)
B_writer = csv.writer(B)
A_headers = [f"x{i}" for i in range(0, 9)]
A_headers.insert(0, "id")
B_headers = [f"x{i}" for i in range(0, 9)]
B_headers.insert(0, "y")
B_headers.insert(0, "id")
A_writer.writerow(A_headers)
B_writer.writerow(B_headers)
for idx, row in enumerate(scaled_data):
    a = row[1:10]
    a = list(a)
    a.insert(0, idx)
    A_writer.writerow(a)
    b = row[9:]
    b[0], b[4] = Y[idx], b[0]
    b = list(b)
    b[0] = int(b[0])
    b.insert(0, idx)
    B_writer.writerow(b)

A.close()
B.close()
